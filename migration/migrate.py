#!/usr/bin/env python3
"""
Migra TrackerVersion4_7.xlsx (Google Sheets exportado) a Supabase/Postgres.

Uso:
    # 1. Sanity check sin tocar la base de datos (recomendado primero):
    python3 migrate.py --xlsx /ruta/TrackerVersion4_7.xlsx --dry-run

    # 2. Crear el esquema (una sola vez):
    psql "$DATABASE_URL" -f schema.sql

    # 3. Cargar los datos:
    python3 migrate.py --xlsx /ruta/TrackerVersion4_7.xlsx --database-url "$DATABASE_URL"

DATABASE_URL es el connection string de Postgres de Supabase (Project Settings
-> Database -> Connection string -> URI). También se puede pasar por variable
de entorno DATABASE_URL en vez de --database-url.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import uuid
from collections import defaultdict

import openpyxl

import lib

# ---------------------------------------------------------------------------
# Clasificación de hojas (ver migration/README.md para el detalle del mapeo)
# ---------------------------------------------------------------------------

SKIP_SHEETS = {
    "Copia de Personal", "Copia de hojaformato 3", "Copia de hojaformato2", "Holtmont",
    "DASHBOARD RAMIRO RODRIGUEZ", "PPCV3",
    "Construcción", "Coordinador HVAC", "Electromecanica", "Limpieza",
}

VENTAS_SHEETS_ORDERED = [
    "ANTONIA_VENTAS",  # maestra: se procesa primero, es la autoridad
    "TERESA GARZA (VENTAS)", "Judith Echavarria (VENTAS)", "Ramiro Rodriguez (VENTAS)",
    "Sebastian Padilla (VENTAS)", "Eduardo Manzanares (VENTAS)", "Angel Salinas (VENTAS)",
    "Miguel Gallardo (VENTAS)", "Eduardo Benitez (VENTAS)", "Cesar Gomez (VENTAS)",
    "Alfonso Correa (VENTAS)", "Juan Jose Sanchez (VENTAS)", "Edgar Lopez (VENTAS)",
    "ALEJANDRO MENDEZ (VENTAS)", "ANTONIA_VENTAS PAPA CALIENTE DE",
    "ANTONIA_VENTAS BANCO DE COTIZAC", "ANTONIA_VENTAS RESUMEN EJECUTIV",
]

WO_SHEETS = {
    "DB_WO_MATERIALES": "wo_materiales",
    "DB_WO_MANO_OBRA": "wo_mano_obra",
    "DB_WO_HERRAMIENTAS": "wo_herramientas",
    "DB_WO_EQUIPOS": "wo_equipos",
    "DB_WO_PROGRAMA": "wo_programa",
}

EXPLICIT_SHEETS = {
    "DB_DIRECTORY", "DB_SITIOS", "DB_PROYECTOS", "PPC_BORRADOR", "DB_BANCO_DATOS",
    "LOG_SISTEMA", "HABITOS_LOG", "KPI_COTIZACIONES", "PPCV4",
}

# folios con generador de secuencia propio (WORKORDER_SEQ, ANTONIA_SEQ, etc.) son
# globales por diseño; un folio puramente numérico corto es un ID legacy que solo
# es único dentro de su propia hoja.
_GLOBAL_FOLIO_PREFIXES = ("PPC-", "AV-", "TG-", "WO-", "SITE-", "PROJ-")


def compute_dedupe_key(folio, source_sheet, row_index):
    if folio:
        if folio.startswith(_GLOBAL_FOLIO_PREFIXES):
            return folio, False
        if folio.replace(".", "", 1).isdigit() and len(folio) >= 10:
            return folio, False  # timestamp-like, ej. Date.now()
        return f"{source_sheet}::{folio}", False
    return f"{source_sheet}::ROW{row_index}", True


# ---------------------------------------------------------------------------
# Fase de extracción + merge en memoria (no toca la DB)
# ---------------------------------------------------------------------------

class MigrationBundle:
    def __init__(self):
        self.people = {}            # nombre_normalizado -> dict
        self.sites = []
        self.projects = []
        self.tasks = {}              # dedupe_key -> dict
        self.task_involucrados = defaultdict(set)  # dedupe_key -> {raw_name}
        self.ppc_borradores = []
        self.quotes = {}             # folio -> dict
        self.banco_datos = []
        self.wo_rows = defaultdict(list)  # table_name -> [rows]
        self.personal_agenda = []
        self.habits_log = []
        self.kpi_cotizaciones = []
        self.system_log = []
        self.warnings = []

    def register_person_field(self, raw):
        """Registra cada nombre de un campo que puede traer uno o varios
        (coma/salto de línea). Nunca registra el string compuesto tal cual
        como si fuera una sola persona -- eso ensucia `people` con filas
        basura tipo 'RAMIRO RODRIGUEZ, ALFONSO CORREA, TERESA GARZA'."""
        for name in lib.split_involucrados(raw):
            self.register_person(name)

    def register_person(self, nombre, departamento=None, tipo_hoja=None):
        if not nombre:
            return
        key = nombre.strip().upper()
        if key not in self.people:
            self.people[key] = {"nombre": key, "departamento": departamento, "tipo_hoja": tipo_hoja}
        else:
            existing = self.people[key]
            if departamento and not existing.get("departamento"):
                existing["departamento"] = departamento
            if tipo_hoja and not existing.get("tipo_hoja"):
                existing["tipo_hoja"] = tipo_hoja


def build_bundle(wb) -> MigrationBundle:
    bundle = MigrationBundle()

    # --- 1. Directorio -------------------------------------------------
    if "DB_DIRECTORY" in wb.sheetnames:
        for p in lib.extract_people(wb["DB_DIRECTORY"]):
            bundle.register_person(p["nombre"], p["departamento"], p["tipo_hoja"])

    # --- 2. Sitios / proyectos ------------------------------------------
    if "DB_SITIOS" in wb.sheetnames:
        bundle.sites = lib.extract_sites(wb["DB_SITIOS"])
    if "DB_PROYECTOS" in wb.sheetnames:
        bundle.projects = lib.extract_projects(wb["DB_PROYECTOS"])

    # --- 3. Trackers personales + PPCV4 ---------------------------------
    tracker_sheet_names = [
        name for name in wb.sheetnames
        if name not in SKIP_SHEETS
        and name not in VENTAS_SHEETS_ORDERED
        and name not in WO_SHEETS
        and name not in EXPLICIT_SHEETS
        and name.strip() != "AGENDA_PERSONAL"
    ]

    def ingest_task(t):
        row_index = t.pop("_row_index")
        raw_inv = t.pop("_involucrados_raw", None)
        key, synthetic = compute_dedupe_key(t.get("folio"), t["source_sheet"], row_index)
        t["dedupe_key"] = key
        t["folio_sintetico"] = synthetic
        if not t.get("folio"):
            t["folio"] = key
        if t.get("assignee_raw"):
            bundle.register_person_field(t["assignee_raw"])
        if key not in bundle.tasks:
            bundle.tasks[key] = t
        else:
            existing = bundle.tasks[key]
            for k, v in t.items():
                if existing.get(k) in (None, "", 0.0) and v not in (None, "", 0.0):
                    existing[k] = v
        for name in lib.split_involucrados(raw_inv):
            bundle.task_involucrados[key].add(name.upper())
            bundle.register_person(name)

    for name in tracker_sheet_names:
        for t in lib.extract_tasks_from_tracker(wb[name], name):
            ingest_task(t)

    if "PPCV4" in wb.sheetnames:
        for t in lib.extract_tasks_from_ppcv4(wb["PPCV4"]):
            ingest_task(t)

    # --- 4. PPC_BORRADOR --------------------------------------------------
    if "PPC_BORRADOR" in wb.sheetnames:
        bundle.ppc_borradores = lib.extract_ppc_borradores(wb["PPC_BORRADOR"])
        for b in bundle.ppc_borradores:
            if b.get("responsable_raw"):
                bundle.register_person_field(b["responsable_raw"])

    # --- 5. Ventas ----------------------------------------------------
    for name in VENTAS_SHEETS_ORDERED:
        if name not in wb.sheetnames:
            continue
        for q in lib.extract_quotes(wb[name], name):
            folio = q["folio"]
            if q.get("vendedor_raw"):
                bundle.register_person_field(q["vendedor_raw"])
            if folio not in bundle.quotes:
                bundle.quotes[folio] = q
            else:
                existing = bundle.quotes[folio]
                for k, v in q.items():
                    if k == "extra":
                        continue
                    if existing.get(k) in (None, "", 0.0) and v not in (None, "", 0.0):
                        existing[k] = v
                if q.get("extra"):
                    merged_extra = dict(existing.get("extra") or {})
                    merged_extra.update(q["extra"])
                    existing["extra"] = merged_extra

    if "DB_BANCO_DATOS" in wb.sheetnames:
        bundle.banco_datos = lib.extract_banco_datos(wb["DB_BANCO_DATOS"])

    # --- 6. Work orders --------------------------------------------------
    for sheet_name, table_name in WO_SHEETS.items():
        if sheet_name in wb.sheetnames:
            bundle.wo_rows[table_name] = lib.extract_wo_table(wb[sheet_name], table_name)

    # --- 7. Agenda / hábitos / KPI / log ---------------------------------
    agenda_name = next((n for n in wb.sheetnames if n.strip() == "AGENDA_PERSONAL"), None)
    if agenda_name:
        bundle.personal_agenda = lib.extract_personal_agenda(wb[agenda_name])
        for a in bundle.personal_agenda:
            if a.get("usuario_raw"):
                bundle.register_person_field(a["usuario_raw"])

    if "HABITOS_LOG" in wb.sheetnames:
        bundle.habits_log = lib.extract_habits_log(wb["HABITOS_LOG"])
        for h in bundle.habits_log:
            if h.get("usuario_raw"):
                bundle.register_person_field(h["usuario_raw"])

    if "KPI_COTIZACIONES" in wb.sheetnames:
        bundle.kpi_cotizaciones = lib.extract_kpi_cotizaciones(wb["KPI_COTIZACIONES"])

    if "LOG_SISTEMA" in wb.sheetnames:
        bundle.system_log = lib.extract_system_log(wb["LOG_SISTEMA"])

    return bundle


def summarize(bundle: MigrationBundle):
    print("=== Resumen de extracción ===")
    print(f"people:              {len(bundle.people)}")
    print(f"sites:                {len(bundle.sites)}")
    print(f"projects:             {len(bundle.projects)}")
    print(f"tasks:                {len(bundle.tasks)}")
    print(f"task_involucrados:    {sum(len(v) for v in bundle.task_involucrados.values())}")
    print(f"ppc_borradores:       {len(bundle.ppc_borradores)}")
    print(f"quotes:               {len(bundle.quotes)}")
    print(f"banco_datos:          {len(bundle.banco_datos)}")
    for table, rows in bundle.wo_rows.items():
        print(f"{table:<21} {len(rows)}")
    print(f"personal_agenda:      {len(bundle.personal_agenda)}")
    print(f"habits_log:           {len(bundle.habits_log)}")
    print(f"kpi_cotizaciones:     {len(bundle.kpi_cotizaciones)}")
    print(f"system_log:           {len(bundle.system_log)}")

    unmatched_assignees = sum(
        1 for t in bundle.tasks.values()
        if t.get("assignee_raw") and t["assignee_raw"] not in bundle.people
    )
    print(f"\ntasks con assignee sin match en people: {unmatched_assignees} (no debería pasar, se auto-registran)")


# ---------------------------------------------------------------------------
# Fase de carga a Postgres
# ---------------------------------------------------------------------------

def load_to_postgres(bundle: MigrationBundle, database_url: str):
    import psycopg2
    import psycopg2.extras as pgx

    pgx.register_uuid()
    conn = psycopg2.connect(database_url)
    conn.autocommit = False
    try:
        with conn.cursor() as cur:
            # 1. people
            people_rows = list(bundle.people.values())
            pgx.execute_values(
                cur,
                """insert into people (nombre, departamento, tipo_hoja) values %s
                   on conflict (nombre) do update set
                       departamento = coalesce(people.departamento, excluded.departamento),
                       tipo_hoja = coalesce(people.tipo_hoja, excluded.tipo_hoja)""",
                [(p["nombre"], p["departamento"], p["tipo_hoja"]) for p in people_rows],
            )
            cur.execute("select nombre, id from people")
            people_ids = {nombre: pid for nombre, pid in cur.fetchall()}
            print(f"people: {len(people_rows)} filas cargadas")

            # 2. sites / projects
            if bundle.sites:
                pgx.execute_values(
                    cur,
                    """insert into sites (id_sitio, nombre, cliente, tipo, estatus, fecha_creacion, creado_por)
                       values %s on conflict (id_sitio) do nothing""",
                    [(s["id_sitio"], s["nombre"], s["cliente"], s["tipo"], s["estatus"],
                      s["fecha_creacion"], s["creado_por"]) for s in bundle.sites],
                )
                print(f"sites: {len(bundle.sites)} filas cargadas")
            if bundle.projects:
                pgx.execute_values(
                    cur,
                    """insert into projects (id_proyecto, id_sitio, nombre_subproyecto, tipo, estatus, fecha_creacion, creado_por)
                       values %s on conflict (id_proyecto) do nothing""",
                    [(p["id_proyecto"], p["id_sitio"], p["nombre_subproyecto"], p["tipo"], p["estatus"],
                      p["fecha_creacion"], p["creado_por"]) for p in bundle.projects],
                )
                print(f"projects: {len(bundle.projects)} filas cargadas")

            # 3. tasks
            task_id_by_key = {}
            task_rows = []
            for key, t in bundle.tasks.items():
                tid = uuid.uuid4()
                task_id_by_key[key] = tid
                task_rows.append((
                    tid, t["folio"], t["dedupe_key"], t["folio_sintetico"],
                    people_ids.get(t.get("assignee_raw")), t.get("assignee_raw"),
                    t.get("departamento"), t.get("fecha_alta"), t.get("hora_alta"),
                    t.get("clasificacion"), t["concepto"], t.get("avance", 0),
                    t.get("fecha_estimada_fin"), t.get("hora_estimada_fin"), t.get("reloj"),
                    t.get("restricciones"), t.get("prioridad"), t.get("riesgos"),
                    t.get("fecha_respuesta"), t.get("correo"), t.get("carpeta"),
                    t.get("cumplimiento"), t.get("comentarios"), t.get("comentarios_semana"),
                    t.get("comentarios_semana_previa"), t.get("status", "PENDIENTE"), t["source_sheet"],
                ))
            pgx.execute_values(
                cur,
                """insert into tasks (
                       id, folio, dedupe_key, folio_sintetico, assignee_id, assignee_raw,
                       departamento, fecha_alta, hora_alta, clasificacion, concepto, avance,
                       fecha_estimada_fin, hora_estimada_fin, reloj, restricciones, prioridad,
                       riesgos, fecha_respuesta, correo, carpeta, cumplimiento, comentarios,
                       comentarios_semana, comentarios_semana_previa, status, source_sheet
                   ) values %s
                   on conflict (dedupe_key) do nothing""",
                task_rows,
            )
            print(f"tasks: {len(task_rows)} filas cargadas")

            # 3b. task_involucrados
            inv_rows = []
            for key, names in bundle.task_involucrados.items():
                tid = task_id_by_key.get(key)
                if not tid:
                    continue
                for name in names:
                    inv_rows.append((tid, people_ids.get(name), name))
            if inv_rows:
                pgx.execute_values(
                    cur,
                    """insert into task_involucrados (task_id, person_id, raw_name) values %s
                       on conflict (task_id, raw_name) do nothing""",
                    inv_rows,
                )
                print(f"task_involucrados: {len(inv_rows)} filas cargadas")

            # 4. ppc_borradores
            if bundle.ppc_borradores:
                pgx.execute_values(
                    cur,
                    """insert into ppc_borradores (
                           especialidad, concepto, responsable_raw, responsable_id, horas, cumplimiento,
                           archivo, comentarios, previos, prioridad, riesgos, restricciones,
                           fecha_respuesta, clasificacion, fecha_alta, ruta_critica, zona, contratista,
                           cuant_requerida, cuant_real, dias_json
                       ) values %s""",
                    [(
                        b["especialidad"], b["concepto"], b["responsable_raw"],
                        people_ids.get(b["responsable_raw"]), b["horas"], b["cumplimiento"],
                        b["archivo"], b["comentarios"], b["previos"], b["prioridad"], b["riesgos"],
                        b["restricciones"], b["fecha_respuesta"], b["clasificacion"], b["fecha_alta"],
                        b["ruta_critica"], b["zona"], b["contratista"], b["cuant_requerida"],
                        b["cuant_real"], b["dias_json"],
                    ) for b in bundle.ppc_borradores],
                )
                print(f"ppc_borradores: {len(bundle.ppc_borradores)} filas cargadas")

            # 5. quotes
            if bundle.quotes:
                pgx.execute_values(
                    cur,
                    """insert into quotes (
                           folio, area, cliente, concepto, clasificacion, vendedor_id, vendedor_raw,
                           f_visita, f_inicio, f_entrega, dias, avance, estatus, comentarios, requisitor,
                           prioridad_cot, info_cliente, f2, cotizacion, timeline, layout, proceso,
                           proceso_log, map_cot, monto, source_sheet, extra
                       ) values %s
                       on conflict (folio) do nothing""",
                    [(
                        q["folio"], q["area"], q["cliente"], q["concepto"], q["clasificacion"],
                        people_ids.get(q.get("vendedor_raw")), q.get("vendedor_raw"),
                        q["f_visita"], q["f_inicio"], q["f_entrega"], q["dias"], q["avance"],
                        q["estatus"], q["comentarios"], q["requisitor"], q["prioridad_cot"],
                        q["info_cliente"], q["f2"], q["cotizacion"], q["timeline"], q["layout"],
                        q["proceso"], q["proceso_log"], q["map_cot"], q["monto"], q["source_sheet"],
                        pgx.Json(q["extra"]) if q.get("extra") else None,
                    ) for q in bundle.quotes.values()],
                )
                print(f"quotes: {len(bundle.quotes)} filas cargadas")

            if bundle.banco_datos:
                pgx.execute_values(
                    cur,
                    """insert into banco_datos (folio, cliente, fecha_inicio, area, concepto, vendedor, estatus, cotizacion, last_update)
                       values %s on conflict (folio) do nothing""",
                    [(b["folio"], b["cliente"], b["fecha_inicio"], b["area"], b["concepto"],
                      b["vendedor"], b["estatus"], b["cotizacion"], b["last_update"]) for b in bundle.banco_datos],
                )
                print(f"banco_datos: {len(bundle.banco_datos)} filas cargadas")

            # 6. work orders
            wo_folios = sorted({row["folio"] for rows in bundle.wo_rows.values() for row in rows})
            if wo_folios:
                pgx.execute_values(
                    cur, "insert into work_orders (folio) values %s on conflict (folio) do nothing",
                    [(f,) for f in wo_folios],
                )
            wo_columns = {
                "wo_materiales": ["folio", "cantidad", "unidad", "tipo", "descripcion", "costo",
                                   "especificacion", "total", "residente", "compras", "controller",
                                   "orden_compra", "pagos", "almacen", "logistica", "residente_obra"],
                "wo_mano_obra": ["folio", "categoria", "salario", "personal", "semanas", "extras",
                                   "nocturno", "fin_semana", "otros", "total"],
                "wo_herramientas": ["folio", "cantidad", "unidad", "descripcion", "costo", "total",
                                      "residente", "controller", "almacen", "logistica", "residente_fin"],
                "wo_equipos": ["folio", "cantidad", "unidad", "tipo", "descripcion", "especificacion",
                                "dias", "horas", "costo", "total"],
                "wo_programa": ["folio", "descripcion", "fecha", "duracion", "unidad_duracion", "unidad",
                                  "cantidad", "precio", "total", "responsable", "seccion", "estatus"],
            }
            for table, rows in bundle.wo_rows.items():
                if not rows:
                    continue
                cols = wo_columns[table]
                pgx.execute_values(
                    cur,
                    f"insert into {table} ({', '.join(cols)}) values %s",
                    [tuple(r.get(c) for c in cols) for r in rows],
                )
                print(f"{table}: {len(rows)} filas cargadas")

            # 7. agenda / hábitos / kpi / log
            if bundle.personal_agenda:
                pgx.execute_values(
                    cur,
                    """insert into personal_agenda (id, usuario_id, usuario_raw, fecha, tipo, hora_inicio, hora_fin, titulo, estatus, detalles)
                       values %s on conflict (id) do nothing""",
                    [(a["id"], people_ids.get(a["usuario_raw"]), a["usuario_raw"], a["fecha"], a["tipo"],
                      a["hora_inicio"], a["hora_fin"], a["titulo"], a["estatus"], a["detalles"]) for a in bundle.personal_agenda],
                )
                print(f"personal_agenda: {len(bundle.personal_agenda)} filas cargadas")

            if bundle.habits_log:
                pgx.execute_values(
                    cur,
                    """insert into habits_log (usuario_id, usuario_raw, habito, meta, log_json, fecha_actualizacion)
                       values %s""",
                    [(people_ids.get(h["usuario_raw"]), h["usuario_raw"], h["habito"], h["meta"],
                      pgx.Json(json.loads(h["log_json"])) if h.get("log_json") else None,
                      h["fecha_actualizacion"]) for h in bundle.habits_log],
                )
                print(f"habits_log: {len(bundle.habits_log)} filas cargadas")

            if bundle.kpi_cotizaciones:
                pgx.execute_values(
                    cur,
                    "insert into kpi_cotizaciones (departamento, total, ganadas, perdidas) values %s",
                    [(k["departamento"], k["total"], k["ganadas"], k["perdidas"]) for k in bundle.kpi_cotizaciones],
                )
                print(f"kpi_cotizaciones: {len(bundle.kpi_cotizaciones)} filas cargadas")

            if bundle.system_log:
                pgx.execute_values(
                    cur,
                    "insert into system_log (fecha_hora, usuario, accion, detalles) values %s",
                    [(s["fecha_hora"], s["usuario"], s["accion"], s["detalles"]) for s in bundle.system_log],
                    page_size=1000,
                )
                print(f"system_log: {len(bundle.system_log)} filas cargadas")

        conn.commit()
        print("\nMigración completada y confirmada (commit).")
    except Exception:
        conn.rollback()
        print("\nERROR: se hizo rollback, no se guardó nada.", file=sys.stderr)
        raise
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--xlsx", required=True, help="Ruta al TrackerVersion4_7.xlsx exportado de Google Sheets")
    parser.add_argument("--database-url", default=os.environ.get("DATABASE_URL"),
                         help="Connection string de Postgres (o variable de entorno DATABASE_URL)")
    parser.add_argument("--dry-run", action="store_true", help="Solo extrae y valida, no toca la base de datos")
    parser.add_argument("--dump-json", help="Si se pasa, escribe el bundle extraído completo a este archivo JSON")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    bundle = build_bundle(wb)
    summarize(bundle)

    if args.dump_json:
        def default(o):
            return str(o)
        with open(args.dump_json, "w", encoding="utf-8") as f:
            json.dump({
                "people": list(bundle.people.values()),
                "tasks": list(bundle.tasks.values()),
                "quotes": list(bundle.quotes.values()),
            }, f, ensure_ascii=False, indent=2, default=default)
        print(f"\nBundle volcado a {args.dump_json}")

    if args.dry_run:
        print("\n--dry-run: no se conectó a la base de datos.")
        return

    if not args.database_url:
        print("\nFalta --database-url (o variable de entorno DATABASE_URL) para cargar los datos.", file=sys.stderr)
        sys.exit(1)

    load_to_postgres(bundle, args.database_url)


if __name__ == "__main__":
    main()
